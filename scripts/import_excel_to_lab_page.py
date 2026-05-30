from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "outputs" / "lab_results_inventory" / "Lab Results Inventory.xlsx"
OUT_FILE = ROOT / "frontend" / "public" / "private" / "liveLabs.json"


NAME_MAP = {
    "GLUCOSE": "Glucose",
    "Glucose": "Glucose",
    "HEMOGLOBIN A1C": "HbA1c",
    "Hemoglobin A1c": "HbA1c",
    "INSULIN": "Fasting Insulin",
    "Insulin": "Fasting Insulin",
    "CREATININE": "Creatinine",
    "Creatinine": "Creatinine",
    "EGFR": "eGFR",
    "eGFR": "eGFR",
    "eGFR NON-AFR. AMERICAN": "eGFR",
    "Glomerular Filtration Rate (eGFR)": "eGFR",
    "ALT": "ALT",
    "ALT (SGPT)": "ALT",
    "AST": "AST",
    "AST (SGOT)": "AST",
    "CHOLESTEROL, TOTAL": "Total Cholesterol",
    "Cholesterol, Total": "Total Cholesterol",
    "Cholesterol": "Total Cholesterol",
    "LDL-CHOLESTEROL": "LDL",
    "LDL Chol Calc (NIH)": "LDL",
    "Cholesterol, LDL (calculated)": "LDL",
    "HDL CHOLESTEROL": "HDL",
    "HDL Cholesterol": "HDL",
    "Cholesterol, HDL": "HDL",
    "TRIGLYCERIDES": "Triglycerides",
    "Triglycerides": "Triglycerides",
    "Triglyceride": "Triglycerides",
    "LDL-P": "LDL-P",
    "Small LDL-P": "Small LDL-P",
    "LP-IR Score": "LP-IR Score",
    "LDL Size": "LDL Size",
    "HDL-P (Total)": "HDL-P",
    "Large HDL-P": "Large HDL-P",
    "WHITE BLOOD CELL COUNT": "WBC",
    "WBC": "WBC",
    "RED BLOOD CELL COUNT": "RBC",
    "RBC": "RBC",
    "HEMOGLOBIN": "Hemoglobin",
    "Hemoglobin": "Hemoglobin",
    "HEMATOCRIT": "Hematocrit",
    "Hematocrit": "Hematocrit",
    "Platelets": "Platelets",
    "PLATELET COUNT": "Platelets",
    "Platelet Count": "Platelets",
    "Neutrophils": "Neutrophils",
    "Neutrophils %": "Neutrophils",
    "Lymphs": "Lymphocytes",
    "Lymphocytes %": "Lymphocytes",
    "Monocytes": "Monocytes",
    "Monocytes %": "Monocytes",
    "Eos": "Eosinophils",
    "Eosinophils %": "Eosinophils",
    "Basos": "Basophils",
    "Basophils %": "Basophils",
    "Testosterone": "Testosterone Total",
    "Testosterone, Total": "Testosterone Total",
    "Free Testosterone(Direct)": "Testosterone Free",
    "Testosterone, Free": "Testosterone Free",
    "C-Reactive Protein, Cardiac": "hs-CRP",
    "Sedimentation Rate-Westergren": "ESR",
    "Specific Gravity": "Specific Gravity",
    "pH": "Urine pH",
    "Protein": "Urine Protein",
    "Ketones": "Urine Ketones",
    "Occult Blood": "Urine Blood",
    "Blood": "Urine Blood",
    "Nitrite, Urine": "Nitrite",
    "WBC Esterase": "Leukocyte Esterase",
}


META = {
    "Glucose": ("Metabolic", "Blood Sugar"),
    "HbA1c": ("Metabolic", "Blood Sugar"),
    "Fasting Insulin": ("Metabolic", "Blood Sugar"),
    "HOMA-IR": ("Metabolic", "Blood Sugar"),
    "Creatinine": ("Metabolic", "Kidney"),
    "eGFR": ("Metabolic", "Kidney"),
    "BUN": ("Metabolic", "Kidney"),
    "ALT": ("Metabolic", "Liver"),
    "AST": ("Metabolic", "Liver"),
    "Alkaline Phosphatase": ("Metabolic", "Liver"),
    "Total Bilirubin": ("Metabolic", "Liver"),
    "Total Cholesterol": ("Lipids", "Standard Panel"),
    "LDL": ("Lipids", "Standard Panel"),
    "HDL": ("Lipids", "Standard Panel"),
    "Triglycerides": ("Lipids", "Standard Panel"),
    "Apolipoprotein B": ("Lipids", "Standard Panel"),
    "LDL-P": ("Lipids", "NMR Profile"),
    "Small LDL-P": ("Lipids", "NMR Profile"),
    "LP-IR Score": ("Lipids", "NMR Profile"),
    "LDL Size": ("Lipids", "NMR Profile"),
    "HDL-P": ("Lipids", "NMR Profile"),
    "Large HDL-P": ("Lipids", "NMR Profile"),
    "WBC": ("CBC", "Cell Counts"),
    "RBC": ("CBC", "Cell Counts"),
    "Hemoglobin": ("CBC", "Cell Counts"),
    "Hematocrit": ("CBC", "Cell Counts"),
    "MCV": ("CBC", "Cell Counts"),
    "MCH": ("CBC", "Cell Counts"),
    "MCHC": ("CBC", "Cell Counts"),
    "RDW": ("CBC", "Cell Counts"),
    "Platelets": ("CBC", "Cell Counts"),
    "Neutrophils": ("CBC", "Differential"),
    "Lymphocytes": ("CBC", "Differential"),
    "Monocytes": ("CBC", "Differential"),
    "Eosinophils": ("CBC", "Differential"),
    "Basophils": ("CBC", "Differential"),
    "Testosterone Total": ("Hormones", "Androgens"),
    "Testosterone Free": ("Hormones", "Androgens"),
    "SHBG": ("Hormones", "Binding Proteins"),
    "hs-CRP": ("Inflammatory", "Cardio Risk"),
    "ESR": ("Inflammatory", "Systemic"),
    "Urine pH": ("Urinalysis", "Chemistry"),
    "Specific Gravity": ("Urinalysis", "Chemistry"),
    "Urine Protein": ("Urinalysis", "Chemistry"),
    "Urine Ketones": ("Urinalysis", "Chemistry"),
    "Urine Blood": ("Urinalysis", "Chemistry"),
    "Urine Glucose": ("Urinalysis", "Chemistry"),
    "Nitrite": ("Urinalysis", "Chemistry"),
    "Leukocyte Esterase": ("Urinalysis", "Chemistry"),
}


def slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


def excel_date(value):
    if value is None:
        return ""
    if hasattr(value, "date"):
        return value.date().isoformat()
    return str(value)[:10]


def canonical_name(test_name: str) -> str:
    cleaned = " ".join(str(test_name).split())
    return NAME_MAP.get(cleaned, NAME_MAP.get(cleaned.upper(), cleaned.title() if cleaned.isupper() else cleaned))


def category_for(name: str, panel: str):
    if name in META:
        return META[name]
    panel_lower = (panel or "").lower()
    if "urinalysis" in panel_lower or name.startswith("Urine"):
        return "Urinalysis", "Chemistry"
    if "nmr" in panel_lower or any(token in name for token in ["LDL-P", "LP-IR", "HDL-P"]):
        return "Lipids", "NMR Profile"
    if "lipid" in panel_lower or "cholesterol" in name.lower() or name in {"LDL", "HDL", "Triglycerides"}:
        return "Lipids", "Standard Panel"
    if "cbc" in panel_lower or name in {"MCV", "MCH", "MCHC", "RDW"}:
        return "CBC", "Cell Counts"
    if "testosterone" in panel_lower:
        return "Hormones", "Androgens"
    if "metabolic" in panel_lower or "cmp" in panel_lower:
        return "Metabolic", "Other"
    return "Metabolic", "Other"


def app_status(status: str, flag: str):
    value = (flag or status or "").lower()
    if value == "high":
        return "high"
    if value == "low":
        return "low"
    if value == "normal":
        return "normal"
    return "unknown"


def display_value(raw, numeric):
    return numeric if numeric is not None else raw


def read_all_results():
    wb = load_workbook(WORKBOOK, data_only=True, read_only=True)
    ws = wb["All Results"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        if record.get("Source File") and record.get("Test Name"):
            records.append(record)
    return records


def result_from_record(record, index):
    test_name = str(record["Test Name"])
    name = canonical_name(test_name)
    date = excel_date(record.get("Collection Date"))
    panel = record.get("Panel / Section") or ""
    if "urinalysis" in panel.lower():
        ua_name_map = {
            "Glucose": "Urine Glucose",
            "Protein": "Urine Protein",
            "Ketones": "Urine Ketones",
            "Occult Blood": "Urine Blood",
            "Blood": "Urine Blood",
            "pH": "Urine pH",
        }
        name = ua_name_map.get(test_name, ua_name_map.get(name, name))
    category, subcategory = category_for(name, panel)
    value = display_value(record.get("Result Value Raw"), record.get("Numeric Value"))
    unit = record.get("Unit") or ""
    source = record.get("Source File") or ""
    page = record.get("Page") or ""
    status = app_status(record.get("Inferred Status"), record.get("Reported Flag"))

    return {
        "id": f"{date}-{slug(source)}-{page}-{slug(name)}-{index}",
        "date": date,
        "provider": record.get("Provider") or "Other",
        "panel": panel or "Unspecified Panel",
        "biomarker": name,
        "displayName": name,
        "value": value,
        "unit": unit,
        "referenceMin": record.get("Reference Min"),
        "referenceMax": record.get("Reference Max"),
        "status": status,
        "category": category,
        "subcategory": subcategory,
        "sourceFile": source,
        "sourcePage": page,
        "notes": "Imported from Lab Results Inventory.xlsx",
        "rawTestName": test_name,
        "rawReferenceRange": record.get("Reference Range Raw") or "",
    }


def add_homa_ir(results):
    grouped = defaultdict(list)
    for result in results:
        grouped[(result["date"], result["provider"], result["sourceFile"])].append(result)

    out = list(results)
    for (date, provider, source), items in grouped.items():
        glucose = next((item for item in items if item["biomarker"] == "Glucose" and isinstance(item["value"], (int, float))), None)
        insulin = next((item for item in items if item["biomarker"] == "Fasting Insulin" and isinstance(item["value"], (int, float))), None)
        if not glucose or not insulin:
            continue
        value = round((glucose["value"] * insulin["value"]) / 405, 2)
        out.append({
            "id": f"{date}-{slug(source)}-homa-ir",
            "date": date,
            "provider": provider,
            "panel": "Calculated from Excel",
            "biomarker": "HOMA-IR",
            "displayName": "HOMA-IR",
            "value": value,
            "unit": "score",
            "referenceMin": 0.5,
            "referenceMax": 2.0,
            "status": "high" if value > 2 else "low" if value < 0.5 else "normal",
            "category": "Metabolic",
            "subcategory": "Blood Sugar",
            "sourceFile": source,
            "sourcePage": "",
            "notes": "Calculated from same-source Excel glucose and fasting insulin rows",
            "rawTestName": "HOMA-IR",
            "rawReferenceRange": "0.5-2.0",
        })
    return out


def main():
    OUT_FILE.write_text("[]\n", encoding="utf-8")
    records = read_all_results()
    results = [result_from_record(record, index) for index, record in enumerate(records, start=1)]
    results = add_homa_ir(results)
    results.sort(key=lambda item: (item["date"], item["provider"], item["sourceFile"], item["biomarker"]))
    OUT_FILE.write_text(json.dumps(results, indent=2) + "\n", encoding="utf-8")
    print(f"Cleared and repopulated {OUT_FILE}")
    print(f"Excel rows imported: {len(records)}")
    print(f"Dashboard rows written: {len(results)}")


if __name__ == "__main__":
    main()
